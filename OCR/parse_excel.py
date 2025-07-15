import pandas as pd
import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
import xlrd  # 添加xlrd库用于处理旧版.xls文件

def parse_excel_with_merged_cells(file_path, active_sheet_only=False):
    """
    解析包含合并单元格的Excel文件，支持多工作表
    
    参数:
        file_path: Excel文件路径
        active_sheet_only: 是否只处理活动工作表(默认False，处理所有工作表)
    
    返回:
        字典: {工作表名: DataFrame}
    """
    # 检查文件类型
    file_ext = os.path.splitext(file_path)[1].lower()
    
    if file_ext == '.xls':
        # 使用xlrd处理.xls文件
        return parse_xls_file(file_path, active_sheet_only)
    else:
        # 使用openpyxl处理.xlsx文件
        return parse_xlsx_file(file_path, active_sheet_only)

def parse_xlsx_file(file_path, active_sheet_only=False):
    """使用openpyxl处理.xlsx文件"""
    wb = load_workbook(file_path)
    sheets_data = {}
    
    # 确定要处理的工作表列表
    sheets_to_process = [wb.active] if active_sheet_only else wb.worksheets
    
    for ws in sheets_to_process:
        # 创建合并单元格映射
        merged_map = {}
        if ws.merged_cells.ranges:
            for merged_range in ws.merged_cells.ranges:
                min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        merged_map[(row, col)] = (min_row, min_col)
        
        # 构建完整数据
        data = []
        for row in ws.iter_rows():
            row_data = []
            for cell in row:
                if (cell.row, cell.column) in merged_map:
                    m_row, m_col = merged_map[(cell.row, cell.column)]
                    if cell.row == m_row and cell.column == m_col:
                        row_data.append(cell.value)
                    else:
                        row_data.append("")  # 标记为合并单元格
                else:
                    row_data.append(cell.value)
            data.append(row_data)
        
        # 创建DataFrame，第一行作为列名
        if data:  # 确保工作表不是空的
            df = pd.DataFrame(data[1:], columns=data[0])
            sheets_data[ws.title] = df
    
    return sheets_data

def parse_xls_file(file_path, active_sheet_only=False):
    """使用xlrd处理.xls文件"""
    wb = xlrd.open_workbook(file_path, formatting_info=True)
    sheets_data = {}
    
    # 确定要处理的工作表列表
    if active_sheet_only:
        sheets_to_process = [wb.sheet_by_index(wb.active_sheet)]
    else:
        sheets_to_process = [wb.sheet_by_index(i) for i in range(wb.nsheets)]
    
    for ws in sheets_to_process:
        # 获取合并单元格信息
        merged_cells = ws.merged_cells
        merged_map = {}
        
        for crange in merged_cells:
            rlo, rhi, clo, chi = crange
            for row in range(rlo, rhi):
                for col in range(clo, chi):
                    merged_map[(row, col)] = (rlo, clo)
        
        # 构建完整数据
        data = []
        for row_idx in range(ws.nrows):
            row_data = []
            for col_idx in range(ws.ncols):
                if (row_idx, col_idx) in merged_map:
                    m_row, m_col = merged_map[(row_idx, col_idx)]
                    if row_idx == m_row and col_idx == m_col:
                        row_data.append(ws.cell_value(row_idx, col_idx))
                    else:
                        row_data.append("")  # 标记为合并单元格
                else:
                    row_data.append(ws.cell_value(row_idx, col_idx))
            data.append(row_data)
        
        # 创建DataFrame，第一行作为列名
        if data:  # 确保工作表不是空的
            df = pd.DataFrame(data[1:], columns=data[0])
            sheets_data[ws.name] = df
    
    return sheets_data

def process_excel_file(input_file, output_folder):
    """处理单个Excel文件并保存为Markdown"""
    print(f"处理文件: {input_file}")
    
    try:
        # 解析Excel文件
        all_sheets = parse_excel_with_merged_cells(input_file)
        
        # 创建输出文件路径
        file_name = os.path.basename(input_file)
        file_stem = os.path.splitext(file_name)[0]
        output_path = os.path.join(output_folder, f"{file_stem}.md")
        
        # 生成Markdown内容
        markdown_text = f"# {file_stem}\n\n"
        for sheet_name, df in all_sheets.items():
            markdown_text += f"## {sheet_name}\n\n"
            markdown_text += df.to_markdown(index=False)
            markdown_text += "\n\n"
        
        # 保存Markdown文件
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(markdown_text)
        
        print(f"Markdown 已保存到：{output_path}")
        return True
    except Exception as e:
        print(f"处理文件 {input_file} 时出错: {e}")
        return False

def process_folder(input_folder, output_folder):
    """处理文件夹中的所有Excel文件"""
    # 确保输入文件夹存在
    if not os.path.exists(input_folder) or not os.path.isdir(input_folder):
        print(f"输入路径不存在或不是文件夹: {input_folder}")
        return
    
    # 创建输出文件夹
    os.makedirs(output_folder, exist_ok=True)
    print(f"输出文件夹: {output_folder}")
    
    # 查找所有Excel文件
    input_path = Path(input_folder)
    excel_files = list(input_path.glob("*.xlsx"))
    excel_files.extend(list(input_path.glob("*.xls")))  # 添加.xls文件
    
    if not excel_files:
        print(f"在 {input_folder} 中未找到Excel文件")
        return
    
    print(f"找到 {len(excel_files)} 个Excel文件")
    
    # 处理每个文件
    successful = 0
    failed = 0
    
    for i, file_path in enumerate(excel_files, 1):
        print(f"处理文件 {i}/{len(excel_files)}: {file_path.name}")
        if process_excel_file(file_path, output_folder):
            successful += 1
        else:
            failed += 1
    
    # 输出统计信息
    print("=" * 50)
    print("处理完成")
    print(f"成功: {successful}/{len(excel_files)}")
    print(f"失败: {failed}/{len(excel_files)}")
    print(f"输出文件保存在: {output_folder}")

if __name__ == "__main__":
    # 设置输入和输出文件夹
    input_folder = "./pdf_files"  # Excel文件所在文件夹
    output_folder = "./output"    # Markdown输出文件夹
    
    # 执行批处理
    process_folder(input_folder, output_folder)
