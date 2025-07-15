#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
文档处理工具模块
提供PDF、Excel、图像等不同格式文档的处理功能
"""

import os
import time
import json
import base64
import logging
import re
from pathlib import Path
from datetime import datetime
from abc import ABC, abstractmethod
from typing import Dict, List, Optional, Union, Any

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# PDF处理相关导入
try:
    from docling.datamodel.base_models import InputFormat
    from docling.datamodel.pipeline_options import PdfPipelineOptions
    from docling.document_converter import DocumentConverter, PdfFormatOption
    from docling_core.types.doc import ImageRefMode
    import warnings
    warnings.filterwarnings("ignore", message="'pin_memory' argument is set as true but not supported on MPS")
    DOCLING_AVAILABLE = True
except ImportError:
    logger.warning("Docling库未安装，PDF处理功能将不可用")
    DOCLING_AVAILABLE = False

# Excel处理相关导入
try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.utils import range_boundaries
    import xlrd
    EXCEL_AVAILABLE = True
except ImportError:
    logger.warning("Excel处理相关库未安装，Excel处理功能将不可用")
    EXCEL_AVAILABLE = False

# 图像处理相关导入
try:
    from openai import OpenAI
    IMAGE_PROCESSING_AVAILABLE = True
except ImportError:
    logger.warning("OpenAI库未安装，图像处理功能将不可用")
    IMAGE_PROCESSING_AVAILABLE = False

# HTML处理相关导入
try:
    import requests
    from bs4 import BeautifulSoup, Comment  # 添加 Comment 导入
    import html2text
    HTML_AVAILABLE = True
except ImportError:
    logger.warning("HTML处理相关库未安装，HTML处理功能将不可用")
    HTML_AVAILABLE = False


class DocumentProcessor(ABC):
    """文档处理器基类"""
    
    def __init__(self, input_path: Union[str, Path], output_path: Union[str, Path]):
        """
        初始化文档处理器
        
        参数:
            input_path: 输入文件或文件夹路径
            output_path: 输出文件或文件夹路径
        """
        self.input_path = Path(input_path)
        self.output_path = Path(output_path)
        
        # 确保输出路径存在
        self.output_path.mkdir(parents=True, exist_ok=True)
    
    @abstractmethod
    def process_file(self, file_path: Path) -> bool:
        """
        处理单个文件
        
        参数:
            file_path: 文件路径
            
        返回:
            bool: 处理是否成功
        """
        pass
    
    def process_folder(self, extensions: List[str], recursive: bool = False) -> Dict[str, int]:
        """
        处理文件夹中的所有指定扩展名的文件
        
        参数:
            extensions: 要处理的文件扩展名列表，如['.pdf', '.xlsx']
            recursive: 是否递归处理子文件夹
            
        返回:
            Dict[str, int]: 处理结果统计，包含成功和失败的文件数
        """
        if not self.input_path.is_dir():
            logger.error(f"输入路径不是文件夹: {self.input_path}")
            return {"success": 0, "fail": 0, "total": 0}
        
        # 查找所有符合条件的文件
        files_to_process = []
        
        if recursive:
            for ext in extensions:
                files_to_process.extend(self.input_path.rglob(f"*{ext}"))
        else:
            for ext in extensions:
                files_to_process.extend(self.input_path.glob(f"*{ext}"))
        
        if not files_to_process:
            logger.warning(f"在 {self.input_path} 中未找到符合条件的文件")
            return {"success": 0, "fail": 0, "total": 0}
        
        logger.info(f"找到 {len(files_to_process)} 个文件需要处理")
        
        # 处理每个文件
        successful = 0
        failed = 0
        
        for i, file_path in enumerate(files_to_process, 1):
            logger.info(f"处理文件 {i}/{len(files_to_process)}: {file_path.name}")
            if self.process_file(file_path):
                successful += 1
            else:
                failed += 1
        
        # 输出统计信息
        logger.info("=" * 50)
        logger.info("处理完成")
        logger.info(f"成功: {successful}/{len(files_to_process)}")
        logger.info(f"失败: {failed}/{len(files_to_process)}")
        logger.info(f"输出文件保存在: {self.output_path}")
        
        return {
            "success": successful,
            "fail": failed,
            "total": len(files_to_process)
        }
    
    def remove_images_from_markdown(self, md_file: Path) -> None:
        """去掉markdown中的图像"""
        with open(md_file, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 移除 ![]() 格式的图片
        content = re.sub(r'!\[.*?\]\(.*?\)', '', content)
        # 移除 <img> 标签
        content = re.sub(r'<img.*?>', '', content)
        
        with open(md_file, 'w', encoding='utf-8') as f:
            f.write(content)


class PdfProcessor(DocumentProcessor):
    """PDF文档处理器"""
    
    def __init__(self, input_path: Union[str, Path], output_path: Union[str, Path], 
                 save_intermediate: bool = True, api_key: Optional[str] = None):
        """
        初始化PDF处理器
        
        参数:
            input_path: 输入文件或文件夹路径
            output_path: 输出文件或文件夹路径
            save_intermediate: 是否保存中间文件
            api_key: 云雾AI API密钥，用于处理扫描文档
        """
        super().__init__(input_path, output_path)
        
        if not DOCLING_AVAILABLE:
            raise ImportError("Docling库未安装，无法处理PDF文件")
        
        self.save_intermediate = save_intermediate
        self.api_key = api_key or os.environ.get("YUNWU_API_KEY")
        self.doc_converter = self._get_docling_converter()
        self.pipeline_options = None
    
    def _get_docling_converter(self):
        """创建并配置Docling转换器"""
        # 创建pipeline配置
        self.pipeline_options = PdfPipelineOptions()
        self.pipeline_options.generate_page_images = True 
        self.pipeline_options.do_ocr = False
        self.pipeline_options.do_table_structure = True
        self.pipeline_options.table_structure_options.do_cell_matching = True
        self.pipeline_options.generate_picture_images = False

        doc_converter = DocumentConverter(
            format_options={
                InputFormat.PDF: PdfFormatOption(pipeline_options=self.pipeline_options)
            }
        )
        return doc_converter
    
    def is_scanned_pdf(self, pdf_path: Path) -> bool:
        """判断PDF文档是否为扫描件"""
        try:
            import fitz  # PyMuPDF
            import numpy as np
            from PIL import Image
            import io
            
            # 打开PDF文件
            doc = fitz.open(pdf_path)
            text_sample = ""
            # Check first few pages for text
            for page_num in range(min(3, len(doc))):
                page = doc.load_page(page_num)
                text_sample += page.get_text("text")
                if len(text_sample.strip()) > 100:  # Found substantial text
                    break
        
            if len(text_sample.strip()) < 50:  # Arbitrary threshold for "very little text"
                logger.info("分析：PDF在前几页中提取到的文本非常少，可能不是纯文本PDF或内容为空。")
                return True
        
            logger.info("分析：PDF初步判断为基于文本的。")
            return False
            
        except Exception as e:
            logger.error(f"检查PDF是否为扫描件时出错: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return True
    
    def process_scanned_page_with_gemini(self, image_path: Path) -> Optional[str]:
        """使用云雾AI API处理扫描页面并返回Markdown内容"""
        if not self.api_key:
            logger.error("未设置云雾API密钥")
            return None
        
        # 读取图像并编码为base64
        def encode_image(image_path):
            with open(image_path, "rb") as image_file:
                return base64.b64encode(image_file.read()).decode('utf-8')
        
        base64_image = encode_image(image_path)
        
        # 创建OpenAI客户端
        client = OpenAI(
            base_url="https://yunwu.ai/v1",
            api_key=self.api_key
        )
        
        try:
            # 调用API
            response = client.chat.completions.create(
                model="gemini-2.0-flash-exp",  # 使用Gemini模型
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {"type": "text", "text": "请提取这个图像中的所有文本内容，并以Markdown格式返回。忽略水印和章，保留原始格式和表格结构。"},
                            {
                                "type": "image_url",
                                "image_url": {
                                    "url": f"data:image/png;base64,{base64_image}"
                                }
                            }
                        ]
                    }
                ],
                temperature=0.1  # 使用较低的温度以获得更确定性的结果
            )
            
            # 提取Markdown内容
            content = response.choices[0].message.content
            return content
        
        except Exception as e:
            logger.error(f"调用云雾API时出错: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return None
    
    def process_scanned_document(self, input_file: Path, pages_dir: Path) -> bool:
        """处理扫描文档，使用Gemini API提取内容"""
        logger.info(f"检测到扫描文档: {input_file.name}，使用Gemini API处理")
            
        # 获取页面总数
        # 假设页面从1开始编号
        page_count = 0
        while True:
            page_file = pages_dir / f"{input_file.stem}-page-{page_count + 1}.png"
            if page_file.exists():
                page_count += 1
            else:
                break
        
        if page_count == 0:
            logger.warning(f"未找到扫描页面图像: {pages_dir}")
            return False
        
        # 创建页面图像列表
        page_images = [pages_dir / f"{input_file.stem}-page-{i}.png" for i in range(1, page_count + 1)]
        
        # 创建单独的MD文件夹
        mds_dir = pages_dir.parent / "mds"
        mds_dir.mkdir(parents=True, exist_ok=True)
        
        try:
            # 处理每一页
            for i, page_image in enumerate(page_images, 1):
                logger.info(f"处理第 {i}/{len(page_images)} 页: {page_image.name}")
                
                # 使用Gemini API处理
                md_content = self.process_scanned_page_with_gemini(page_image)
                    
                if md_content:
                    # 保存单独的页面MD文件
                    page_md_path = mds_dir / f"{input_file.stem}-page-{i}.md"
                    with page_md_path.open("w", encoding="utf-8") as page_md_file:
                        page_md_file.write(md_content)
                    logger.info(f"已保存页面 {i} 的Markdown到: {page_md_path}")
                else:
                    logger.warning(f"无法提取页面内容: {page_image.name}")
                
            logger.info(f"单独的页面Markdown文件保存在: {mds_dir}")
            
            # 合并MD文件
            self.combine_md_files(input_file, mds_dir)
            
            return True
        except Exception as e:
            logger.error(f"处理扫描文档时出错: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return False
    
    def combine_md_files(self, input_file: Path, mds_dir: Path) -> bool:
        """将mds文件夹中的MD文件按编号组合成一个完整的文件"""
        logger.info(f"开始合并MD文件: {input_file.stem}")
        
        # 创建输出Markdown文件
        md_path = self.output_path / f"{input_file.stem}.md"
        
        try:
            # 查找所有相关的MD文件
            md_files = list(mds_dir.glob(f"{input_file.stem}-page-*.md"))
            
            if not md_files:
                logger.warning(f"未找到MD文件: {mds_dir}")
                return False
            
            # 按文件名中的页码数字排序
            md_files.sort(key=lambda x: int(x.stem.split('-page-')[1]))
            
            # 合并内容
            all_content = [f"# {input_file.stem}\n\n"]
            
            for i, md_file in enumerate(md_files, 1):
                logger.info(f"合并第 {i} 页: {md_file.name}")
                
                try:
                    with md_file.open("r", encoding="utf-8") as f:
                        content = f.read()
                        all_content.append(f"## 第 {i} 页\n\n")
                        all_content.append(content)
                        all_content.append("\n\n")
                except Exception as e:
                    logger.error(f"读取文件 {md_file} 时出错: {e}")
                    # 继续处理其他文件
            
            # 写入合并后的文件
            with md_path.open("w", encoding="utf-8") as f:
                f.write("".join(all_content))
            
            # 检查文件是否成功写入
            if md_path.exists() and md_path.stat().st_size > 0:
                logger.info(f"成功合并MD文件到: {md_path} (文件大小: {md_path.stat().st_size} 字节)")
                return True
            else:
                logger.error(f"合并后的MD文件为空或不存在: {md_path}")
                return False
                
        except Exception as e:
            logger.error(f"合并MD文件时出错: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return False
    def process_file(self, file_path: Path) -> bool:
        """处理单个PDF文件并保存结果"""
        logger.info(f"开始处理文件: {file_path.name}")
        start_time = time.time()
        
        # 首先判断输入文件是否为扫描件
        is_scanned = self.is_scanned_pdf(file_path)
        if is_scanned:
            logger.info(f"检测到扫描PDF文档: {file_path.name}")
        
        pages_dir = None  # 初始化页面目录变量

        try:
            # 转换文档
            conv_result = self.doc_converter.convert(str(file_path))
            
            # 处理时间
            processing_time = time.time() - start_time
            logger.info(f"文档转换完成，耗时: {processing_time:.2f}秒")
            
            # 保存中间文件
            if self.save_intermediate:
                intermediate_dir = self.output_path / f"{file_path.stem}_details"
                intermediate_dir.mkdir(parents=True, exist_ok=True)
                
                # 保存页面图像
                if conv_result.document.pages and hasattr(next(iter(conv_result.document.pages.values())), 'image'):
                    pages_dir = intermediate_dir / "pages"
                    pages_dir.mkdir(parents=True, exist_ok=True)
                    for page_no, page in conv_result.document.pages.items():
                        if page.image and page.image.pil_image:
                            page_image_path = pages_dir / f"{file_path.stem}-page-{page.page_no}.png"
                            with page_image_path.open("wb") as fp:
                                page.image.pil_image.save(fp, format="PNG")
                    logger.info(f"已保存页面图像到: {pages_dir}")
                
                # 保存JSON格式
                json_path = intermediate_dir / f"{file_path.stem}.json"
                with json_path.open("w", encoding="utf-8") as fp:
                    fp.write(json.dumps(conv_result.document.export_to_dict(), ensure_ascii=False))
                logger.info(f"已保存JSON输出到: {json_path}")
                
                # 保存提取的图片
                if hasattr(self.pipeline_options, 'generate_picture_images') and self.pipeline_options.generate_picture_images and conv_result.document.pictures:
                    pictures_dir = intermediate_dir / "pictures"
                    pictures_dir.mkdir(parents=True, exist_ok=True)
                    for pic_idx, picture_item in enumerate(conv_result.document.pictures):
                        if picture_item.image and picture_item.image.pil_image:
                            pic_path = pictures_dir / f"{file_path.stem}_pic_{pic_idx}.png"
                            picture_item.image.pil_image.save(pic_path)
                    logger.info(f"已保存提取的图片到: {pictures_dir}")
            
            # 如果不是扫描件，保存普通Markdown文件
            if not is_scanned:
                # 获取Markdown内容
                md_content = conv_result.document.export_to_markdown(image_mode=ImageRefMode.EMBEDDED)
                
                md_path = self.output_path / f"{file_path.stem}.md"
                with md_path.open("w", encoding="utf-8") as fp:
                    fp.write(md_content)
                
                # 移除图像
                self.remove_images_from_markdown(md_path)
                logger.info(f"成功导出Markdown到: {md_path}")
            else:
                logger.info(f"检测到扫描文档，跳过普通Markdown导出")
            
            # 如果是扫描文档且已生成页面图像，则使用Gemini API处理
            if is_scanned and pages_dir and pages_dir.exists():
                self.process_scanned_document(file_path, pages_dir)
            
            return True
        
        except Exception as e:
            logger.error(f"处理文件 {file_path.name} 时出错: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return False


class ExcelProcessor(DocumentProcessor):
    """Excel文档处理器"""
    
    def __init__(self, input_path: Union[str, Path], output_path: Union[str, Path]):
        """
        初始化Excel处理器
        
        参数:
            input_path: 输入文件或文件夹路径
            output_path: 输出文件或文件夹路径
        """
        super().__init__(input_path, output_path)
        
        if not EXCEL_AVAILABLE:
            raise ImportError("Excel处理相关库未安装，无法处理Excel文件")
    
    def parse_excel_with_merged_cells(self, file_path: Path, active_sheet_only: bool = False) -> Dict[str, pd.DataFrame]:
        """
        解析包含合并单元格的Excel文件，支持多工作表
        
        参数:
            file_path: Excel文件路径
            active_sheet_only: 是否只处理活动工作表(默认False，处理所有工作表)
        
        返回:
            字典: {工作表名: DataFrame}
        """
        # 检查文件类型
        file_ext = file_path.suffix.lower()
        
        if file_ext == '.xls':
            # 使用xlrd处理.xls文件
            return self._parse_xls_file(file_path, active_sheet_only)
        else:
            # 使用openpyxl处理.xlsx文件
            return self._parse_xlsx_file(file_path, active_sheet_only)
    
    def _parse_xlsx_file(self, file_path: Path, active_sheet_only: bool = False) -> Dict[str, pd.DataFrame]:
        """使用openpyxl处理.xlsx文件"""
        logger.info(f"使用openpyxl处理.xlsx文件: {file_path.name}")
        wb = load_workbook(file_path)
        sheets_data = {}
        
        # 确定要处理的工作表列表
        sheets_to_process = [wb.active] if active_sheet_only else wb.worksheets
        
        for ws in sheets_to_process:
            logger.info(f"处理工作表: {ws.title}")
            # 创建合并单元格映射
            merged_map = {}
            if ws.merged_cells.ranges:
                for merged_range in ws.merged_cells.ranges:
                    min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
                    for row in range(min_row, max_row + 1):
                        for col in range(min_col, max_col + 1):
                            merged_map[(row, col)] = (min_row, min_col)
            
            # 构建完整数据
            data = []
            for row in ws.iter_rows():
                row_data = []
                for cell in row:
                    if (cell.row, cell.column) in merged_map:
                        m_row, m_col = merged_map[(cell.row, cell.column)]
                        if cell.row == m_row and cell.column == m_col:
                            row_data.append(cell.value)
                        else:
                            row_data.append("")  # 标记为合并单元格
                    else:
                        row_data.append(cell.value)
                data.append(row_data)
            
            # 创建DataFrame，第一行作为列名
            if data:  # 确保工作表不是空的
                if len(data) > 1:
                    df = pd.DataFrame(data[1:], columns=data[0])
                else:
                    # 如果只有一行（列名），创建空DataFrame但保留列名
                    df = pd.DataFrame(columns=data[0])
                sheets_data[ws.title] = df
        
        return sheets_data
    
    def _parse_xls_file(self, file_path: Path, active_sheet_only: bool = False) -> Dict[str, pd.DataFrame]:
        """使用xlrd处理.xls文件"""
        logger.info(f"使用xlrd处理.xls文件: {file_path.name}")
        wb = xlrd.open_workbook(file_path, formatting_info=True)
        sheets_data = {}
        
        # 确定要处理的工作表列表
        if active_sheet_only:
            sheets_to_process = [wb.sheet_by_index(wb.active_sheet)]
        else:
            sheets_to_process = [wb.sheet_by_index(i) for i in range(wb.nsheets)]
        
        for ws in sheets_to_process:
            logger.info(f"处理工作表: {ws.name}")
            # 获取合并单元格信息
            merged_cells = ws.merged_cells
            merged_map = {}
            
            for crange in merged_cells:
                rlo, rhi, clo, chi = crange
                for row in range(rlo, rhi):
                    for col in range(clo, chi):
                        merged_map[(row, col)] = (rlo, clo)
            
            # 构建完整数据
            data = []
            for row_idx in range(ws.nrows):
                row_data = []
                for col_idx in range(ws.ncols):
                    if (row_idx, col_idx) in merged_map:
                        m_row, m_col = merged_map[(row_idx, col_idx)]
                        if row_idx == m_row and col_idx == m_col:
                            row_data.append(ws.cell_value(row_idx, col_idx))
                        else:
                            row_data.append("")  # 标记为合并单元格
                    else:
                        row_data.append(ws.cell_value(row_idx, col_idx))
                data.append(row_data)
            
            # 创建DataFrame，第一行作为列名
            if data:  # 确保工作表不是空的
                if len(data) > 1:
                    df = pd.DataFrame(data[1:], columns=data[0])
                else:
                    # 如果只有一行（列名），创建空DataFrame但保留列名
                    df = pd.DataFrame(columns=data[0])
                sheets_data[ws.name] = df
        
        return sheets_data
    
    def parse_excel_with_pandas(self, file_path: Path, sheet_name=None) -> Dict[str, pd.DataFrame]:
        """
        使用pandas直接解析Excel文件（不处理合并单元格）
        
        参数:
            file_path: Excel文件路径
            sheet_name: 要处理的工作表名称，None表示所有工作表
            
        返回:
            字典: {工作表名: DataFrame}
        """
        logger.info(f"使用pandas解析Excel文件: {file_path.name}")
        try:
            # 检查文件类型
            file_ext = file_path.suffix.lower()
            
            # 根据文件类型选择不同的处理方式
            if file_ext == '.xlsx':
                # 对于.xlsx文件，使用openpyxl引擎并设置data_only=True以获取计算结果而非公式
                excel_file = pd.ExcelFile(file_path, engine='openpyxl', engine_kwargs={'data_only': True})
            elif file_ext == '.xls':
                # 对于.xls文件，使用xlrd引擎
                excel_file = pd.ExcelFile(file_path, engine='xlrd')
            else:
                # 其他类型使用默认引擎
                excel_file = pd.ExcelFile(file_path)
            
            # 确定要处理的工作表
            sheets_to_process = sheet_name or excel_file.sheet_names
            if isinstance(sheets_to_process, str):
                sheets_to_process = [sheets_to_process]
            
            # 解析每个工作表
            result = {}
            for sheet in sheets_to_process:
                if sheet in excel_file.sheet_names:
                    df = excel_file.parse(sheet)
                    result[sheet] = df
                else:
                    logger.warning(f"工作表 '{sheet}' 不存在于文件 {file_path.name} 中")
            
            return result
            
        except Exception as e:
            logger.error(f"使用pandas解析Excel文件时出错: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return {}
    
    def convert_to_markdown(self, sheets_data: Dict[str, pd.DataFrame], title: str = None) -> str:
        """
        将Excel数据转换为Markdown格式
        
        参数:
            sheets_data: 工作表数据字典 {工作表名: DataFrame}
            title: Markdown文档标题，默认为None
            
        返回:
            str: Markdown格式的文本
        """
        markdown_text = []
        
        # 添加标题
        if title:
            markdown_text.append(f"# {title}\n\n")
        
        # 处理每个工作表
        for sheet_name, df in sheets_data.items():
            markdown_text.append(f"## {sheet_name}\n\n")
            
            # 转换DataFrame为Markdown表格
            try:
                # 检查DataFrame是否为空
                if df.empty:
                    markdown_text.append("*此工作表为空*\n\n")
                else:
                    # 使用to_markdown方法转换为Markdown表格
                    table_md = df.to_markdown(index=False)
                    markdown_text.append(f"{table_md}\n\n")
            except Exception as e:
                logger.error(f"转换工作表 '{sheet_name}' 为Markdown时出错: {e}")
                markdown_text.append(f"*转换工作表 '{sheet_name}' 时出错*\n\n")
        
        return "".join(markdown_text)
    
    def process_file(self, file_path: Path) -> bool:
        """处理单个Excel文件并保存结果"""
        logger.info(f"开始处理Excel文件: {file_path.name}")
        start_time = time.time()
        
        try:
            # 解析Excel文件，处理合并单元格
            sheets_data = self.parse_excel_with_merged_cells(file_path)
            
            if not sheets_data:
                logger.warning(f"未能从文件 {file_path.name} 中提取任何数据")
                return False
            
            # 处理时间
            processing_time = time.time() - start_time
            logger.info(f"Excel解析完成，耗时: {processing_time:.2f}秒")
            
            # 转换为Markdown
            file_stem = file_path.stem
            markdown_text = self.convert_to_markdown(sheets_data, title=file_stem)
            
            # 保存Markdown文件
            md_path = self.output_path / f"{file_stem}.md"
            with md_path.open("w", encoding="utf-8") as f:
                f.write(markdown_text)
            
            # 检查文件是否成功写入
            if md_path.exists() and md_path.stat().st_size > 0:
                logger.info(f"成功导出Markdown到: {md_path} (文件大小: {md_path.stat().st_size} 字节)")
                return True
            else:
                logger.error(f"Markdown文件为空或不存在: {md_path}")
                return False
            
        except Exception as e:
            logger.error(f"处理文件 {file_path.name} 时出错: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return False


class ImageProcessor(DocumentProcessor):
    """图像文档处理器，使用云雾AI的Gemini模型提取图像中的文本"""
    
    def __init__(self, input_path: Union[str, Path], output_path: Union[str, Path], 
                 api_key: Optional[str] = None, model: str = "gemini-2.0-flash-thinking-exp-01-21"):
        """
        初始化图像处理器
        
        参数:
            input_path: 输入文件或文件夹路径
            output_path: 输出文件或文件夹路径
            api_key: 云雾AI API密钥
            model: 使用的Gemini模型名称
        """
        super().__init__(input_path, output_path)
        
        if not IMAGE_PROCESSING_AVAILABLE:
            raise ImportError("OpenAI库未安装，无法处理图像文件")
        
        self.api_key = api_key or os.environ.get("YUNWU_API_KEY")
        if not self.api_key:
            logger.warning("未设置云雾AI API密钥，图像文本提取将不可用")
        
        self.model = model
    
    def encode_image(self, image_path: Path) -> str:
        """将图像编码为base64格式"""
        with open(image_path, "rb") as image_file:
            return base64.b64encode(image_file.read()).decode('utf-8')
    
    def get_mime_type(self, file_path: Path) -> str:
        """根据文件扩展名获取MIME类型"""
        # 获取文件扩展名
        ext = file_path.suffix.lower().lstrip('.')
        
        # 确定MIME类型
        mime_types = {
            'jpg': 'image/jpeg',
            'jpeg': 'image/jpeg',
            'png': 'image/png',
            'gif': 'image/gif',
            'webp': 'image/webp',
            'bmp': 'image/bmp',
            'tiff': 'image/tiff',
            'tif': 'image/tiff'
        }
        
        return mime_types.get(ext, 'image/jpeg')  # 默认为jpeg
    
    def extract_text_from_image(self, image_path: Path, prompt: Optional[str] = None) -> Optional[str]:
        """
        从图像中提取文本
        
        参数:
            image_path: 图像文件路径
            prompt: 自定义提示词，如果为None则使用默认提示词
            
        返回:
            提取的文本内容，如果提取失败则返回None
        """
        if not self.api_key:
            logger.error("未设置云雾AI API密钥，无法提取图像文本")
            return None
        
        # 检查文件是否存在
        if not image_path.exists():
            logger.error(f"文件不存在: {image_path}")
            return None
        
        # 获取MIME类型
        mime_type = self.get_mime_type(image_path)
        
        # 编码图像
        try:
            base64_image = self.encode_image(image_path)
        except Exception as e:
            logger.error(f"无法读取或编码图像: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return None
        
        # 创建OpenAI客户端
        client = OpenAI(
            base_url="https://yunwu.ai/v1",
            api_key=self.api_key
        )
        
        # 默认提示词
        if prompt is None:
            prompt = "请提取这个图像中的所有文本内容，并以Markdown格式返回。忽略水印和章，保留原始格式和表格结构。"
        
        # 调用API
        try:
            logger.info(f"正在处理图像: {image_path.name}")
            start_time = time.time()
            
            response = client.chat.completions.create(
                model=self.model,
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {"type": "text", "text": prompt},
                            {
                                "type": "image_url",
                                "image_url": {
                                    "url": f"data:{mime_type};base64,{base64_image}"
                                }
                            }
                        ]
                    }
                ],
                temperature=0.1  # 使用较低的温度以获得更确定性的结果
            )
            
            processing_time = time.time() - start_time
            logger.info(f"处理完成，耗时: {processing_time:.2f}秒")
            
            # 提取内容
            content = response.choices[0].message.content
            return content
            
        except Exception as e:
            logger.error(f"调用云雾AI API时出错: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return None
    
    def process_file(self, file_path: Path) -> bool:
        """处理单个图像文件并保存结果"""
        logger.info(f"开始处理图像文件: {file_path.name}")
        
        try:
            # 从图像中提取文本
            md_content = self.extract_text_from_image(file_path)
            
            if not md_content:
                logger.warning(f"未能从图像 {file_path.name} 中提取文本")
                return False
            
            # 保存Markdown文件
            md_path = self.output_path / f"{file_path.stem}.md"
            with md_path.open("w", encoding="utf-8") as f:
                f.write(md_content)
            
            # 检查文件是否成功写入
            if md_path.exists() and md_path.stat().st_size > 0:
                logger.info(f"成功导出Markdown到: {md_path} (文件大小: {md_path.stat().st_size} 字节)")
                return True
            else:
                logger.error(f"Markdown文件为空或不存在: {md_path}")
                return False
            
        except Exception as e:
            logger.error(f"处理文件 {file_path.name} 时出错: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return False
    
    def process_folder(self, extensions: List[str] = None, recursive: bool = False) -> Dict[str, int]:
        """
        处理文件夹中的所有图像文件
        
        参数:
            extensions: 要处理的文件扩展名列表，如['.jpg', '.png']
            recursive: 是否递归处理子文件夹
            
        返回:
            Dict[str, int]: 处理结果统计，包含成功和失败的文件数
        """
        # 如果未指定扩展名，使用默认的图像扩展名
        if not extensions:
            extensions = ['.jpg', '.jpeg', '.png', '.gif', '.webp', '.bmp', '.tiff', '.tif']
        
        logger.info(f"开始处理图像文件夹，支持的扩展名: {extensions}")
        
        # 调用父类的process_folder方法
        results = super().process_folder(extensions=extensions, recursive=recursive)
        
        return results
    
    def batch_process_images(self, image_paths: List[Path]) -> Dict[str, bool]:
        """
        批量处理指定的图像文件
        
        参数:
            image_paths: 图像文件路径列表
            
        返回:
            Dict[str, bool]: 处理结果，{文件名: 是否成功}
        """
        results = {}
        
        for image_path in image_paths:
            logger.info(f"处理图像: {image_path.name}")
            success = self.process_file(image_path)
            results[image_path.name] = success
        
        # 输出统计信息
        successful = sum(1 for success in results.values() if success)
        failed = len(results) - successful
        
        logger.info("=" * 50)
        logger.info("批量处理完成")
        logger.info(f"成功: {successful}/{len(results)}")
        logger.info(f"失败: {failed}/{len(results)}")
        
        return results


class HTMLProcessor(DocumentProcessor):
    """HTML文档处理器，将HTML文件转换为Markdown格式"""
    
    def __init__(self, input_path: Union[str, Path], output_path: Union[str, Path]):
        """
        初始化HTML处理器
        
        参数:
            input_path: 输入文件或文件夹路径
            output_path: 输出文件或文件夹路径
        """
        super().__init__(input_path, output_path)
        
        if not HTML_AVAILABLE:
            raise ImportError("HTML处理相关库未安装，无法处理HTML文件")
        
        # 配置html2text转换器
        self.html2text_handler = html2text.HTML2Text()
        self.html2text_handler.ignore_links = False
        self.html2text_handler.ignore_images = False
        self.html2text_handler.ignore_tables = False
        self.html2text_handler.body_width = 0  # 不自动换行
        self.html2text_handler.unicode_snob = True  # 使用Unicode字符
        self.html2text_handler.single_line_break = True  # 单行换行
        self.html2text_handler.wrap_links = False  # 不换行链接
        self.html2text_handler.mark_code = True  # 标记代码块
    
    def extract_title(self, soup: BeautifulSoup) -> str:
        """从HTML中提取标题"""
        title_tag = soup.find('title')
        if title_tag and title_tag.string:
            return title_tag.string.strip()
        
        # 如果没有title标签，尝试查找h1标签
        h1_tag = soup.find('h1')
        if h1_tag and h1_tag.string:
            return h1_tag.string.strip()
        
        return "无标题"
    
    def clean_html(self, html_content: str) -> str:
        """清理HTML内容，移除不需要的元素"""
        soup = BeautifulSoup(html_content, 'html.parser')
        
        # 移除脚本和样式
        for script in soup(["script", "style"]):
            script.decompose()
        
        # 移除注释
        for comment in soup.find_all(string=lambda text: isinstance(text, Comment)):
            comment.extract()
        
        # 可以添加更多清理逻辑，如移除导航栏、页脚等
        
        return str(soup)
    
    def html_to_markdown(self, html_content: str) -> str:
        """将HTML转换为Markdown"""
        # 清理HTML
        clean_content = self.clean_html(html_content)
        
        # 使用html2text转换为Markdown
        markdown = self.html2text_handler.handle(clean_content)
        
        return markdown
    
    def process_file(self, file_path: Path) -> bool:
        """处理单个HTML文件并保存结果"""
        logger.info(f"开始处理HTML文件: {file_path.name}")
        
        try:
            # 读取HTML文件
            with open(file_path, 'r', encoding='utf-8') as f:
                html_content = f.read()
            
            # 解析HTML
            soup = BeautifulSoup(html_content, 'html.parser')
            
            # 提取标题
            title = self.extract_title(soup)
            logger.info(f"提取到标题: {title}")
            
            # 转换为Markdown
            markdown_content = self.html_to_markdown(html_content)
            
            # 添加标题
            if title and not markdown_content.startswith('# '):
                markdown_content = f"# {title}\n\n{markdown_content}"
            
            # 保存Markdown文件
            md_path = self.output_path / f"{file_path.stem}.md"
            with md_path.open("w", encoding="utf-8") as f:
                f.write(markdown_content)
            self.remove_images_from_markdown(md_path)
            # 检查文件是否成功写入
            if md_path.exists() and md_path.stat().st_size > 0:
                logger.info(f"成功导出Markdown到: {md_path} (文件大小: {md_path.stat().st_size} 字节)")
                return True
            else:
                logger.error(f"Markdown文件为空或不存在: {md_path}")
                return False
            
        except Exception as e:
            logger.error(f"处理文件 {file_path.name} 时出错: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return False
